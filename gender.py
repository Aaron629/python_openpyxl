import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill,Color
from openpyxl.formula.translate import Translator 
from django.db.models.functions import TruncYear as Year
from django.db.models import Value as V
from django.db.models.functions import Coalesce

years=[106,107,108,109,110]
gender=[1,2,3]
json_gender={
  "io": 1,
  "n": [
    [564,9,260],
    [572,9,319],
    [584,7,352],
    [567,6,303],
    [632,5,277]
  ],
  "gender": [[1,3,2]]
}
wb = openpyxl.load_workbook("C://Users//許皓倫//github//gender_0.xlsx")
active_sheet = wb.worksheets[0]
active_sheet.insert_cols(4, (len(years)-1)*2)

for col_a in active_sheet['A4:A6']:
    for cell in col_a:
      cell.font = Font(name='微軟正黑體',size=12)

for row in active_sheet['B2:C6']:
  for cell in row:
    cell.font = Font(name='微軟正黑體',size=12)
      
source_range = active_sheet['B2:C7']

# initial_range = active_sheet['B2:C6']
for cell in source_range[0]:  # 確保選擇的是第7列的所有單元格
    cell.font = Font(name='微軟正黑體',size=12,)
a_columns=[['B','C'],['D','E'],['F','G'],['H','I'],['J','K'],['L','M'],['N','O'],['P','Q']]

total_row_4=[]
total_row_5=[]
total_row_6=[]
for i in range(len(years)):
  total_row_4.append(f'{a_columns[i][0]}4')
  total_row_5.append(f'{a_columns[i][0]}5')
  total_row_6.append(f'{a_columns[i][0]}6')
  print(json_gender['n'][i][0]) 
  # print(json_gender['n'][i][1])
  # print(json_gender['n'][i][2])
  for rows in range(len(source_range)):
    for cols in range(len(source_range[0])):
      active_sheet[f'{a_columns[i][0]}2']=years[i]
      active_sheet[f'{a_columns[i][0]}3']='人數'
      active_sheet[f'{a_columns[i][1]}3']='百分比' 
      source_cell = source_range[rows][cols]
      destination_cell = active_sheet.cell(row=2+rows,column=4+i*2+cols)
      destination_cell.font = Font(
        name='微軟正黑體',
        size=source_cell.font.size,
        bold=source_cell.font.bold,
        vertAlign=source_cell.font.vertAlign,
        color=source_cell.font.color
      )
      destination_cell.alignment = Alignment(horizontal='center', vertical='center')
      destination_cell.border = Border(
        left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color) if source_cell.border.left else None,
        right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color) if source_cell.border.right else None,
        top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color) if source_cell.border.top else None,
        bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color) if source_cell.border.bottom else None,
        diagonal=Side(style=source_cell.border.diagonal.style, color=source_cell.border.diagonal.color) if source_cell.border.diagonal else None,
        diagonal_direction=source_cell.border.diagonal_direction,
        outline=source_cell.border.outline,
        vertical=Side(style=source_cell.border.vertical.style, color=source_cell.border.vertical.color) if source_cell.border.vertical else None,
        horizontal=Side(style=source_cell.border.horizontal.style, color=source_cell.border.horizontal.color) if source_cell.border.horizontal else None
      )
      destination_cell.fill = PatternFill(
        fill_type=source_cell.fill.fill_type,
        start_color=source_cell.fill.start_color,
        end_color=source_cell.fill.end_color
      )
      # destination_cell.value = source_cell.value
      # if(rows>3):
      for j in range(len(gender)+1):
        if j<len(gender):
          active_sheet.cell(row=4+j,column=2+i*2).value=json_gender['n'][i][j] #106人數
          active_sheet.cell(row=4+j,column=2+i*2).font=Font(name='微軟正黑體',size=12) #106人數
          active_sheet.cell(row=4+j,column=2+i*2).alignment = Alignment(horizontal='center', vertical='center') 
          active_sheet.cell(row=4+j,column=3+i*2).value=f'={a_columns[i][0]}{4+j}/{a_columns[i][0]}7' #106百分比
          active_sheet.cell(row=4+j,column=3+i*2).number_format = '0.00%'
          active_sheet.cell(row=4+j,column=3+i*2).alignment = Alignment(horizontal='center', vertical='center')
          active_sheet.cell(row=4+j,column=4+i*2).alignment = Alignment(horizontal='center', vertical='center')
          active_sheet.cell(row=4+j,column=5+i*2).value=f'={a_columns[i+1][0]}{4+j}/{a_columns[i+1][0]}7' #個別百分比
          active_sheet.cell(row=4+j,column=5+i*2).alignment = Alignment(horizontal='center', vertical='center')  
          active_sheet.cell(row=4+j,column=5+i*2).number_format = '0.00%'
        else:
          active_sheet.cell(row=4+j,column=2+i*2).value=f'=SUM({a_columns[i][0]}4:{a_columns[i][0]}6)' #106合計
          active_sheet.cell(row=4+j,column=2+i*2).alignment = Alignment(horizontal='center', vertical='center')
          active_sheet.cell(row=4+j,column=3+i*2).value=f'=SUM({a_columns[i][1]}4:{a_columns[i][1]}6)' #107合計
          active_sheet.cell(row=4+j,column=3+i*2).alignment = Alignment(horizontal='center', vertical='center')
          active_sheet.cell(row=4+j,column=3+i*2).number_format = '0.00%'
          active_sheet.cell(row=4+j,column=5+i*2).number_format = '0.00%'
        active_sheet.cell(row=4,column=4+i*2).value=f'=SUM({",".join(total_row_4)})'
        active_sheet.cell(row=5,column=4+i*2).value=f'=SUM({",".join(total_row_5)})'
        active_sheet.cell(row=6,column=4+i*2).value=f'=SUM({",".join(total_row_6)})'
        active_sheet.cell(row=7,column=4+i*2).value=f'=SUM({a_columns[i+1][0]}4:{a_columns[i+1][0]}6)' #總計合計
        active_sheet.cell(row=7,column=4+i*2).alignment = Alignment(horizontal='center', vertical='center')
        active_sheet.cell(row=7,column=5+i*2).value=f'=SUM({a_columns[i+1][1]}4:{a_columns[i+1][1]}6)' #百分比合計
        active_sheet.cell(row=7,column=5+i*2).alignment = Alignment(horizontal='center', vertical='center')
  active_sheet.merge_cells(start_row=2, start_column=2+i*2, end_row=2, end_column=3+i*2) #合併年分    
# 設置 A7 的字體顏色為紅色
print(total_row_4,active_sheet.cell(row=4,column=4+i*2).value)
red_color = Color(rgb="FFFF0000")  # 使用 RGB 顏色代碼設置紅色
active_sheet['A7'].font = Font(name='微軟正黑體',size=12,color=red_color)

# 確保第七行的其他單元格字體顏色與 A7 一致
for cell in active_sheet[f'B7:{a_columns[i+1][1]}7'][0]:  # 確保選擇的是第7列的所有單元格
    cell.font = Font(name='微軟正黑體',size=12,color=active_sheet['A7'].font.color)
    
active_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
active_sheet['A1'].font =Font(size=16,bold=True)
active_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5+i*2) #合併標題欄
active_sheet.merge_cells(start_row=2, start_column=4+i*2, end_row=3, end_column=4+i*2) #合併總計
active_sheet.merge_cells(start_row=2, start_column=5+i*2, end_row=3, end_column=5+i*2 )#合併百分比

wb.save("C://Users//許皓倫//github//gender_1.xlsx")
